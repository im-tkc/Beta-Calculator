#!/usr/bin/env python

import sys
import os
from YahooToExcel import YahooToExcel
from openpyxl import load_workbook
import argparse
from openpyxl.chart.trendline import Trendline
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

def export(ticker, frequency, filename, workbook=None): # 1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo
    json_file = "{0}.json".format(filename)
    output_file = "{0}.xlsx".format(filename)
    
    y2e = YahooToExcel()
    workbook = y2e.run(json_file, ticker, frequency, output_file, workbook)

    return workbook
    
def copy_data(workbook, output_file, index_ticker, company_ticker):
    sheets_list = workbook.sheetnames
    beta_output = sheets_list[0]
    company_ticker = sheets_list[1]
    market_ticker = sheets_list[2]
    beta_worksheet = workbook[beta_output]
    company_worksheet = workbook[company_ticker]
    market_worksheet = workbook[market_ticker]
    basic_offset = 3
    
    company_change_len = len(company_worksheet['H'])
    market_end_change_len = len(market_worksheet['H'])
    market_start_change_len = len(market_worksheet['H']) - company_change_len + basic_offset # include header and baseline
    
    beta_worksheet["A1"] = '{0} % change'.format(company_ticker)
    beta_worksheet["B1"] = '{0} % change'.format(index_ticker)
    for i in range(basic_offset, company_change_len+1):
        beta_worksheet["A{0}".format(i-1)] = company_worksheet["H{0}".format(i)].value
        beta_worksheet["B{0}".format(i-1)] = market_worksheet["H{0}".format(i + market_start_change_len - basic_offset)].value
        
    workbook.save(output_file)
    return workbook
    
    
def plot_graph(workbook, output_file, index_ticker, company_ticker):
    sheets_list = workbook.sheetnames
    beta_output = sheets_list[0]
    company_ticker = sheets_list[1]
    market_ticker = sheets_list[2]
    beta_worksheet = workbook[beta_output]
    company_change_len = len(beta_worksheet['A'])

    chart = ScatterChart()
    chart.title = "Beta"
    chart.style = 13
    chart.x_axis.title = '{0} % change'.format(index_ticker)
    chart.y_axis.title = '{0} % change'.format(company_ticker)
    
    xvalues = Reference(beta_worksheet, min_col=2, min_row = 2, max_row = company_change_len)
    yvalues = Reference(beta_worksheet, min_col=1, min_row = 2, max_row = company_change_len)
    
    series = Series(yvalues, xvalues, title_from_data=False)
    series.graphicalProperties.line.noFill = True
    series.marker.symbol = "circle"
    series.marker.size = "2"
    series.trendline = Trendline(dispRSqr=True, trendlineType='linear', dispEq=True)
    
    chart.series.append(series)
    beta_worksheet.add_chart(chart, "F5")
    workbook.save(output_file)
        
def main(argv):
    filename = None
    
    if len(sys.argv) == 1:
        print("Beta calculator")
        print("    Shortcut tips: run.py -m <index_ticker from Yahoo Finance> -c <company ticker from Yahoo Finance> -f <frequency> -o <output file name>")
        
        index_ticker = input("Market ticker: ")
        company_ticker = input("Company ticker: ")
        frequency = input("Frequency (1d, 5d, 1wk, 1mo, 3mo): ")
        filename = input("Output filename: ")
        
    elif len(sys.argv) == 9:
        parser = argparse.ArgumentParser(description="Beta calculator")
        parser.add_argument('-m', action='store', dest="index_ticker")
        parser.add_argument('-s', action='store', dest="company_ticker")
        parser.add_argument('-f', action='store', dest="frequency")
        parser.add_argument('-o', action='store', dest="output")
        
        result = parser.parse_args()
        index_ticker = result.index_ticker
        company_ticker = result.company_ticker
        frequency = result.frequency
        filename = result.output
    else:
        print("Beta calculator")
        print("    Shortcut tips: run.py -m <index_ticker> -c <company ticker> -f <frequency> -o <output file name>")
        sys.exit(1)
        
    workbook = export(index_ticker, frequency, filename)
    workbook = export(company_ticker, frequency, filename, workbook)
    
    output_file = "{0}.xlsx".format(filename)
    workbook = load_workbook(output_file)
    workbook = YahooToExcel().calculate_beta(workbook, output_file)
    workbook = copy_data(workbook, output_file, index_ticker, company_ticker)
    plot_graph(workbook, output_file, index_ticker, company_ticker)
    
    os.remove("{0}.json".format(filename))
    
if __name__ == '__main__':
    main(sys.argv)