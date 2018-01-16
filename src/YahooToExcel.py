#!/usr/bin/env python

import sys
import json
import urllib.request
from datetime import datetime, timedelta
import time
from openpyxl import Workbook
import numpy
import PricePoint

class YahooToExcel(object):
    @staticmethod
    def download(ticker, end_time, frequency, filename):
        url = "https://query1.finance.yahoo.com/v8/finance/chart/{0}?symbol={1}&period1=-{2}&period2={3}&interval={4}&includePrePost=true&events=div|split|earn&corsDomain=finance.yahoo.com".format(ticker, ticker, end_time, end_time, frequency)
        urllib.request.urlretrieve(url, filename)
    
    @staticmethod
    def to_datetime(unixtime):
        # exception takes place in if it is less than 86400 (within a day of epoch time)
        if unixtime < 86400:
            adjusted_datetime = datetime(1970, 1, 1) + timedelta(seconds=unixtime)
            return adjusted_datetime.strftime('%Y-%m-%d %H:%M:%S')
        else:
            return datetime.fromtimestamp(unixtime).strftime('%Y-%m-%d %H:%M:%S')
    
    @staticmethod    
    def export_excel(json_data, filename, ticker, output_file, workbook=None):
        if workbook == None:
            workbook = Workbook()
        worksheet = workbook.create_sheet(ticker, 0)
        worksheet["A1"] = "date"
        worksheet["B1"] = "open"
        worksheet["C1"] = "high"
        worksheet["D1"] = "low"
        worksheet["E1"] = "close"
        worksheet["F1"] = "adj close"
        worksheet["G1"] = "volume"
        worksheet["H1"] = "% change"
        
        i = 0
        unixtime = json_data['chart']['result'][0]['timestamp'][i]
        low_price = json_data['chart']['result'][0]['indicators']['quote'][0]['low'][i]
        close_price = json_data['chart']['result'][0]['indicators']['quote'][0]['close'][i]
        high_price = json_data['chart']['result'][0]['indicators']['quote'][0]['high'][i]
        open_price = json_data['chart']['result'][0]['indicators']['quote'][0]['open'][i]
        volume = json_data['chart']['result'][0]['indicators']['quote'][0]['volume'][i]
        adj_close_price = json_data['chart']['result'][0]['indicators']['adjclose'][0]['adjclose'][i]
        percent_change = None
        is_valid_input = (open_price is not None) and (high_price is not None) and (low_price is not None) and (close_price is not None) and (adj_close_price is not None)
        
        date_time = YahooToExcel.to_datetime(unixtime)
        if is_valid_input:
            worksheet["A{0}".format(i+2)] = date_time
            worksheet["B{0}".format(i+2)] = open_price
            worksheet["C{0}".format(i+2)] = high_price
            worksheet["D{0}".format(i+2)] = low_price
            worksheet["E{0}".format(i+2)] = close_price
            worksheet["F{0}".format(i+2)] = adj_close_price
            worksheet["G{0}".format(i+2)] = volume
            worksheet["H{0}".format(i+2)] = percent_change
        
        offset = 0
        for i in range(1, len(json_data['chart']['result'][0]['timestamp'])):
            unixtime = json_data['chart']['result'][0]['timestamp'][i]
            low_price = json_data['chart']['result'][0]['indicators']['quote'][0]['low'][i]
            close_price = json_data['chart']['result'][0]['indicators']['quote'][0]['close'][i]
            high_price = json_data['chart']['result'][0]['indicators']['quote'][0]['high'][i]
            open_price = json_data['chart']['result'][0]['indicators']['quote'][0]['open'][i]
            volume = json_data['chart']['result'][0]['indicators']['quote'][0]['volume'][i]
            adj_close_price = json_data['chart']['result'][0]['indicators']['adjclose'][0]['adjclose'][i]
            prev_adj_close = json_data['chart']['result'][0]['indicators']['adjclose'][0]['adjclose'][i-1]
            counter = 1
            while True:
                if prev_adj_close is None:
                    prev_adj_close = json_data['chart']['result'][0]['indicators']['adjclose'][0]['adjclose'][i-counter]
                    counter+=1
                    offset -=1
                else:
                    break

            date_time = YahooToExcel.to_datetime(unixtime)
            is_valid_input = (open_price is not None) and (high_price is not None) and (low_price is not None) and (close_price is not None) and (adj_close_price is not None)
            
            if is_valid_input:
                percent_change = (adj_close_price - prev_adj_close) / prev_adj_close
                worksheet["A{0}".format(i+2+offset)] = date_time
                worksheet["B{0}".format(i+2+offset)] = open_price
                worksheet["C{0}".format(i+2+offset)] = high_price
                worksheet["D{0}".format(i+2+offset)] = low_price
                worksheet["E{0}".format(i+2+offset)] = close_price
                worksheet["F{0}".format(i+2+offset)] = adj_close_price
                worksheet["G{0}".format(i+2+offset)] = volume
                worksheet["H{0}".format(i+2+offset)] = percent_change

        workbook.save(output_file)
        return workbook
        
    @staticmethod    
    def load_to_memory(json_data, filename, output_file):
        pricePoints = []
        for i in range(1, len(json_data['chart']['result'][0]['timestamp'])):
            unixtime = json_data['chart']['result'][0]['timestamp'][i]
            low_price = json_data['chart']['result'][0]['indicators']['quote'][0]['low'][i]
            close_price = json_data['chart']['result'][0]['indicators']['quote'][0]['close'][i]
            high_price = json_data['chart']['result'][0]['indicators']['quote'][0]['high'][i]
            open_price = json_data['chart']['result'][0]['indicators']['quote'][0]['open'][i]
            volume = json_data['chart']['result'][0]['indicators']['quote'][0]['volume'][i]
            adj_close_price = json_data['chart']['result'][0]['indicators']['adjclose'][0]['adjclose'][i]
            prev_adj_close = json_data['chart']['result'][0]['indicators']['adjclose'][0]['adjclose'][i-1]
            percent_change = (adj_close_price - prev_adj_close) / prev_adj_close
            
            date_time = YahooToExcel.to_datetime(unixtime)
            is_valid_input = (open_price is not None) and (high_price is not None) and (low_price is not None) and (close_price is not None) and (adj_close_price is not None)
            
            if is_valid_input:
                pp = PricePoint()
                pp.timestamp = unixtime
                pp.open_price = open_price
                pp.close_price = close_price
                pp.high_price = high_price
                pp.low_price = low_price
                pp.adj_close_price = adj_close_price
                pp.volume = volume
                pp.percent_change = percent_change
                pricePoints.append(pp)
    
    @staticmethod
    def calculate_beta(workbook, output_file):
        sheets_list = workbook.sheetnames
        company_ticker = sheets_list[0]
        market_ticker = sheets_list[1]
        company_worksheet = workbook[company_ticker]
        market_worksheet = workbook[market_ticker]
        
        company_change_len = len(company_worksheet['H'])
        market_end_change_len = len(market_worksheet['H'])
        market_start_change_len = len(market_worksheet['H']) - company_change_len + 3 # include header and baseline
        
        worksheet = workbook.create_sheet("beta result", 0)
        worksheet["A1"] = "Regression beta"
        worksheet["A2"] = "Adjusted beta"
        worksheet["B1"] = "=_xlfn.COVARIANCE.S('{0}'!H3:H{1}, '{2}'!H{3}:H{4})/_xlfn.VAR.S('{5}'!H{6}:H{7})".format(
                                company_ticker, company_change_len, 
                                market_ticker, market_start_change_len, market_end_change_len, 
                                market_ticker, market_start_change_len, market_end_change_len
                            )
        worksheet["B2"] = "=B1 * 2 / 3"
        
                                
        workbook.save(output_file)
        return workbook    
    @staticmethod
    def get_json(filename):
        raw_data = None
        with open(filename, 'rb') as file:
            raw_data=file.read()
        return json.loads(raw_data)
    
    @staticmethod
    def run(filename, ticker, frequency, output_file, workbook=None):
        current_unixtime = time.mktime(datetime.now().timetuple())
        YahooToExcel.download(ticker, int(current_unixtime), frequency, filename)
        json_data = YahooToExcel.get_json(filename)
        workbook = YahooToExcel.export_excel(json_data, filename, ticker, output_file, workbook)

        return workbook
        
        